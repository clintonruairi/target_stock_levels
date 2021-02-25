import requests
import json
import urllib
from openpyxl import load_workbook


def search(upc, store_id_number):
    availability_endpoint = f"https://redsky.target.com/redsky_aggregations/v1/web/plp_search_v1?key=ff457966e64d5e877fdbad070f276d18ecec4a01&channel=WEB&keyword={upc}&page=%2Fs%2F{upc}&pricing_store_id={store_id_number}&store_ids={store_id_number}&visitor_id=01776CACF61B0201A5EDFF43711F1193"
    response = requests.get(availability_endpoint)
    json_response = json.loads(response.text)
    with open(f'{upc}-{store_id_number}-price.json', 'w') as file:
        json.dump(json_response, file)
    if json_response.get("data").get("search").get("search_response").get("typed_metadata").get("total_results") == 0:
        print(f'Product {upc} not available at store {store_id_number}')
        return
    title = json_response.get("data").get("search").get("products")[0].get("item").get("product_description").get("title")
    tcin = json_response.get("data").get("search").get("products")[0].get("tcin")
    price = json_response.get("data").get("search").get("products")[0].get("price").get("current_retail")
    in_stock_endpoint = f'https://redsky.target.com/redsky_aggregations/v1/web/pdp_fulfillment_v1?key=ff457966e64d5e877fdbad070f276d18ecec4a01&tcin={tcin}&store_id={store_id_number}&store_positions_store_id={store_id_number}&has_store_positions_store_id=true&pricing_store_id={store_id_number}'
    in_stock_response = requests.get(in_stock_endpoint)
    in_stock_json = json.loads(in_stock_response.text)
    quantity = in_stock_json.get('data').get('product').get('fulfillment').get('store_options')[0].get('location_available_to_promise_quantity')
    in_stock_at_location = in_stock_json.get('data').get('product').get('fulfillment').get('store_options')[0].get('in_store_only').get('availability_status')
    with open(f'{upc}-{store_id_number}-in-stock.json', 'w') as file:
        json.dump(in_stock_json, file)
    print(f'Title: {title}')
    print(f'Price: {price}, UPC: {upc}, STORE: {store_id_number}')
    print(f'Promised Quantity In Store: {quantity}')
    print(f'Availability Status: {in_stock_at_location}')
    print('\n')


def excel_searcher():
    wb = load_workbook('Products.xlsx')
    sheet = wb['Sheet1']
    end_row = sheet.max_row
    start_row = sheet.min_row + 1
    upcs = []
    for row in sheet[f'A{start_row}': f'A{end_row}']:
        for cell in row:
            upcs.append(str(cell.value))
    store_ids = []
    for row in sheet[f'B{start_row}': f'B{end_row}']:
        for cell in row:
            store_ids.append(str(cell.value))
    counter = 0
    for upc in upcs:
        search(upc, store_ids[counter])
        counter += 1

excel_searcher()


    


